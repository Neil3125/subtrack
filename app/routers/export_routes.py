"""Export routes for generating reports."""
from fastapi import APIRouter, Depends, Response
from sqlalchemy.orm import Session
from datetime import date, timedelta
from collections import defaultdict
from app.database import get_db
from app.models import Subscription, Customer, Category
from io import BytesIO
import csv

router = APIRouter()


@router.get("/export/subscriptions/excel")
async def export_subscriptions_excel(db: Session = Depends(get_db)):
    """Export subscriptions to Excel format."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Subscriptions"
        
        # Headers
        headers = ["ID", "Vendor", "Plan", "Cost", "Currency", "Billing Cycle", "Status", 
                   "Customer", "Customer Email", "Customer Country", "Category", "Subscription Country", "Next Renewal", "Notes"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Get data
        subscriptions = db.query(Subscription).all()
        
        for sub in subscriptions:
            ws.append([
                sub.id,
                sub.vendor_name,
                sub.plan_name or "",
                float(sub.cost),
                sub.currency,
                sub.billing_cycle.value,
                sub.status.value,
                sub.customer.name if sub.customer else "",
                sub.customer.email if sub.customer and sub.customer.email else "",
                sub.customer.country if sub.customer and sub.customer.country else "",
                sub.category.name if sub.category else "",
                sub.country or "",
                sub.next_renewal_date.isoformat() if sub.next_renewal_date else "",
                sub.notes or ""
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=subscriptions_{date.today().isoformat()}.xlsx"}
        )
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return await export_subscriptions_csv(db)


@router.get("/export/subscriptions/csv")
async def export_subscriptions_csv(db: Session = Depends(get_db)):
    """Export subscriptions to CSV format."""
    output = BytesIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow(["ID", "Vendor", "Plan", "Cost", "Currency", "Billing Cycle", "Status", 
                     "Customer", "Customer Email", "Customer Country", "Category", "Subscription Country", "Next Renewal", "Notes"])
    
    # Data
    subscriptions = db.query(Subscription).all()
    for sub in subscriptions:
        writer.writerow([
            sub.id,
            sub.vendor_name,
            sub.plan_name or "",
            sub.cost,
            sub.currency,
            sub.billing_cycle.value,
            sub.status.value,
            sub.customer.name if sub.customer else "",
            sub.customer.email if sub.customer and sub.customer.email else "",
            sub.customer.country if sub.customer and sub.customer.country else "",
            sub.category.name if sub.category else "",
            sub.country or "",
            sub.next_renewal_date.isoformat() if sub.next_renewal_date else "",
            sub.notes or ""
        ])
    
    output.seek(0)
    return Response(
        content=output.getvalue().decode('utf-8'),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=subscriptions_{date.today().isoformat()}.csv"}
    )


@router.get("/export/analytics/excel")
async def export_analytics_excel(db: Session = Depends(get_db)):
    """Export analytics report to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Get data
        active_subs = db.query(Subscription).filter(Subscription.status == "active").all()
        all_subs = db.query(Subscription).all()
        
        total_cost = sum(s.cost for s in active_subs)
        
        # Summary data
        ws_summary.append(["SubTrack Analytics Report"])
        ws_summary.append([f"Generated: {date.today().isoformat()}"])
        ws_summary.append([])
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Active Subscriptions", len(active_subs)])
        ws_summary.append(["Total Monthly Cost", f"${total_cost:.2f}"])
        ws_summary.append(["Average Cost per Subscription", f"${total_cost/len(active_subs):.2f}" if active_subs else "$0.00"])
        ws_summary.append(["Total Subscriptions (All Status)", len(all_subs)])
        
        # Style summary
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_summary['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        
        for row in ws_summary['A4:B8']:
            for cell in row:
                if cell.column == 1:
                    cell.font = Font(bold=True)
        
        # By Category sheet
        ws_category = wb.create_sheet("By Category")
        ws_category.append(["Category", "Count", "Total Cost"])
        
        categories = db.query(Category).all()
        for cat in categories:
            cat_subs = [s for s in active_subs if s.category_id == cat.id]
            cat_cost = sum(s.cost for s in cat_subs)
            ws_category.append([cat.name, len(cat_subs), cat_cost])
        
        # Style headers
        for cell in ws_category[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # By Vendor sheet
        ws_vendor = wb.create_sheet("By Vendor")
        ws_vendor.append(["Vendor", "Count", "Total Cost"])
        
        from collections import defaultdict
        vendor_stats = defaultdict(lambda: {"count": 0, "total": 0})
        for sub in active_subs:
            vendor_stats[sub.vendor_name]["count"] += 1
            vendor_stats[sub.vendor_name]["total"] += sub.cost
        
        for vendor, stats in sorted(vendor_stats.items(), key=lambda x: x[1]["total"], reverse=True):
            ws_vendor.append([vendor, stats["count"], stats["total"]])
        
        # Style headers
        for cell in ws_vendor[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=analytics_report_{date.today().isoformat()}.xlsx"}
        )
    except ImportError:
        # Fallback to CSV
        output = BytesIO()
        writer = csv.writer(output)
        
        writer.writerow(["SubTrack Analytics Report"])
        writer.writerow([f"Generated: {date.today().isoformat()}"])
        writer.writerow([])
        
        active_subs = db.query(Subscription).filter(Subscription.status == "active").all()
        total_cost = sum(s.cost for s in active_subs)
        
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Active Subscriptions", len(active_subs)])
        writer.writerow(["Total Monthly Cost", f"${total_cost:.2f}"])
        
        output.seek(0)
        return Response(
            content=output.getvalue().decode('utf-8'),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=analytics_report_{date.today().isoformat()}.csv"}
        )


@router.get("/export/outstanding/excel")
async def export_outstanding_excel(db: Session = Depends(get_db)):
    """Export outstanding (overdue and expiring soon) subscriptions to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        today = date.today()
        threshold_date = today + timedelta(days=30)
        
        # Overdue subscriptions sheet
        ws_overdue = wb.active
        ws_overdue.title = "Overdue"
        
        headers = ["ID", "Vendor", "Plan", "Cost", "Currency", "Customer", "Customer Email", "Customer Country", "Category", 
                   "Subscription Country", "Next Renewal", "Days Overdue"]
        ws_overdue.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws_overdue[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        overdue_subs = db.query(Subscription).filter(
            Subscription.status == "active",
            Subscription.next_renewal_date < today
        ).all()
        
        total_overdue_cost = 0
        for sub in overdue_subs:
            days_overdue = (today - sub.next_renewal_date).days
            total_overdue_cost += sub.cost
            ws_overdue.append([
                sub.id,
                sub.vendor_name,
                sub.plan_name or "",
                float(sub.cost),
                sub.currency,
                sub.customer.name if sub.customer else "",
                sub.customer.email if sub.customer and sub.customer.email else "",
                sub.customer.country if sub.customer and sub.customer.country else "",
                sub.category.name if sub.category else "",
                sub.country or "Not specified",
                sub.next_renewal_date.isoformat() if sub.next_renewal_date else "",
                days_overdue
            ])
        
        # Expiring soon sheet
        ws_expiring = wb.create_sheet("Expiring Soon (30 days)")
        headers_expiring = ["ID", "Vendor", "Plan", "Cost", "Currency", "Customer", "Customer Email", "Customer Country", "Category", 
                           "Subscription Country", "Next Renewal", "Days Until Renewal"]
        ws_expiring.append(headers_expiring)
        
        header_fill_warning = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        for cell in ws_expiring[1]:
            cell.fill = header_fill_warning
            cell.font = header_font
        
        expiring_subs = db.query(Subscription).filter(
            Subscription.status == "active",
            Subscription.next_renewal_date >= today,
            Subscription.next_renewal_date <= threshold_date
        ).all()
        
        total_expiring_cost = 0
        for sub in expiring_subs:
            days_until = (sub.next_renewal_date - today).days
            total_expiring_cost += sub.cost
            ws_expiring.append([
                sub.id,
                sub.vendor_name,
                sub.plan_name or "",
                float(sub.cost),
                sub.currency,
                sub.customer.name if sub.customer else "",
                sub.customer.email if sub.customer and sub.customer.email else "",
                sub.customer.country if sub.customer and sub.customer.country else "",
                sub.category.name if sub.category else "",
                sub.country or "Not specified",
                sub.next_renewal_date.isoformat() if sub.next_renewal_date else "",
                days_until
            ])
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Outstanding Subscriptions Report"])
        ws_summary.append([f"Generated: {today.isoformat()}"])
        ws_summary.append([])
        ws_summary.append(["Metric", "Count", "Total Cost"])
        ws_summary.append(["Overdue Subscriptions", len(overdue_subs), f"${total_overdue_cost:.2f}"])
        ws_summary.append(["Expiring in 30 Days", len(expiring_subs), f"${total_expiring_cost:.2f}"])
        ws_summary.append(["Total Outstanding", len(overdue_subs) + len(expiring_subs), f"${total_overdue_cost + total_expiring_cost:.2f}"])
        
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        # Auto-adjust column widths for all sheets
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=outstanding_subscriptions_{today.isoformat()}.xlsx"}
        )
    except ImportError:
        # Fallback to CSV
        output = BytesIO()
        writer = csv.writer(output)
        today = date.today()
        
        writer.writerow(["Outstanding Subscriptions Report"])
        writer.writerow([f"Generated: {today.isoformat()}"])
        writer.writerow([])
        writer.writerow(["ID", "Vendor", "Plan", "Cost", "Currency", "Customer", "Category", "Country", "Next Renewal", "Status"])
        
        overdue_subs = db.query(Subscription).filter(
            Subscription.status == "active",
            Subscription.next_renewal_date < today
        ).all()
        
        for sub in overdue_subs:
            writer.writerow([sub.id, sub.vendor_name, sub.plan_name or "", sub.cost, sub.currency,
                           sub.customer.name if sub.customer else "", sub.category.name if sub.category else "",
                           sub.country or "Not specified", sub.next_renewal_date.isoformat() if sub.next_renewal_date else "", "OVERDUE"])
        
        output.seek(0)
        return Response(
            content=output.getvalue().decode('utf-8'),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=outstanding_subscriptions_{today.isoformat()}.csv"}
        )


@router.get("/export/country-count/excel")
async def export_country_count_excel(db: Session = Depends(get_db)):
    """Export subscription count by country to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.chart import BarChart, Reference
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Subscriptions by Country"
        
        headers = ["Country", "Active Subscriptions", "Total Subscriptions", "Percentage of Active"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Get all subscriptions grouped by country
        all_subs = db.query(Subscription).all()
        active_subs = [s for s in all_subs if s.status.value == "active"]
        
        country_stats = defaultdict(lambda: {"active": 0, "total": 0})
        for sub in all_subs:
            country = sub.country or "Not Specified"
            country_stats[country]["total"] += 1
            if sub.status.value == "active":
                country_stats[country]["active"] += 1
        
        total_active = len(active_subs)
        
        # Sort by active count descending
        for country, stats in sorted(country_stats.items(), key=lambda x: x[1]["active"], reverse=True):
            percentage = (stats["active"] / total_active * 100) if total_active > 0 else 0
            ws.append([country, stats["active"], stats["total"], f"{percentage:.1f}%"])
        
        # Add totals row
        ws.append([])
        ws.append(["TOTAL", total_active, len(all_subs), "100%"])
        ws[ws.max_row][0].font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=subscriptions_by_country_{date.today().isoformat()}.xlsx"}
        )
    except ImportError:
        output = BytesIO()
        writer = csv.writer(output)
        
        writer.writerow(["Subscriptions by Country Report"])
        writer.writerow([f"Generated: {date.today().isoformat()}"])
        writer.writerow([])
        writer.writerow(["Country", "Active Subscriptions", "Total Subscriptions"])
        
        all_subs = db.query(Subscription).all()
        country_stats = defaultdict(lambda: {"active": 0, "total": 0})
        for sub in all_subs:
            country = sub.country or "Not Specified"
            country_stats[country]["total"] += 1
            if sub.status.value == "active":
                country_stats[country]["active"] += 1
        
        for country, stats in sorted(country_stats.items(), key=lambda x: x[1]["active"], reverse=True):
            writer.writerow([country, stats["active"], stats["total"]])
        
        output.seek(0)
        return Response(
            content=output.getvalue().decode('utf-8'),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=subscriptions_by_country_{date.today().isoformat()}.csv"}
        )


@router.get("/export/country-revenue/excel")
async def export_country_revenue_excel(db: Session = Depends(get_db)):
    """Export subscription revenue by country to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revenue by Country"
        
        headers = ["Country", "Active Revenue", "Total Revenue", "Active Subscriptions", "Avg Revenue per Subscription"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Get all subscriptions grouped by country
        all_subs = db.query(Subscription).all()
        
        country_stats = defaultdict(lambda: {"active_revenue": 0, "total_revenue": 0, "active_count": 0, "total_count": 0})
        for sub in all_subs:
            country = sub.country or "Not Specified"
            country_stats[country]["total_revenue"] += sub.cost
            country_stats[country]["total_count"] += 1
            if sub.status.value == "active":
                country_stats[country]["active_revenue"] += sub.cost
                country_stats[country]["active_count"] += 1
        
        total_active_revenue = sum(s["active_revenue"] for s in country_stats.values())
        total_revenue = sum(s["total_revenue"] for s in country_stats.values())
        
        # Sort by active revenue descending
        for country, stats in sorted(country_stats.items(), key=lambda x: x[1]["active_revenue"], reverse=True):
            avg_revenue = stats["active_revenue"] / stats["active_count"] if stats["active_count"] > 0 else 0
            ws.append([
                country, 
                f"${stats['active_revenue']:.2f}", 
                f"${stats['total_revenue']:.2f}",
                stats["active_count"],
                f"${avg_revenue:.2f}"
            ])
        
        # Add totals row
        ws.append([])
        total_active_count = sum(s["active_count"] for s in country_stats.values())
        total_avg = total_active_revenue / total_active_count if total_active_count > 0 else 0
        ws.append(["TOTAL", f"${total_active_revenue:.2f}", f"${total_revenue:.2f}", total_active_count, f"${total_avg:.2f}"])
        ws[ws.max_row][0].font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=revenue_by_country_{date.today().isoformat()}.xlsx"}
        )
    except ImportError:
        output = BytesIO()
        writer = csv.writer(output)
        
        writer.writerow(["Revenue by Country Report"])
        writer.writerow([f"Generated: {date.today().isoformat()}"])
        writer.writerow([])
        writer.writerow(["Country", "Active Revenue", "Total Revenue", "Active Subscriptions"])
        
        all_subs = db.query(Subscription).all()
        country_stats = defaultdict(lambda: {"active_revenue": 0, "total_revenue": 0, "active_count": 0})
        for sub in all_subs:
            country = sub.country or "Not Specified"
            country_stats[country]["total_revenue"] += sub.cost
            if sub.status.value == "active":
                country_stats[country]["active_revenue"] += sub.cost
                country_stats[country]["active_count"] += 1
        
        for country, stats in sorted(country_stats.items(), key=lambda x: x[1]["active_revenue"], reverse=True):
            writer.writerow([country, f"${stats['active_revenue']:.2f}", f"${stats['total_revenue']:.2f}", stats["active_count"]])
        
        output.seek(0)
        return Response(
            content=output.getvalue().decode('utf-8'),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=revenue_by_country_{date.today().isoformat()}.csv"}
        )


# ==================== Data Persistence Endpoints ====================

@router.get("/export/data-backup")
async def export_data_backup(db: Session = Depends(get_db)):
    """
    Export all data as JSON for backup purposes.
    This can be used to manually backup your data.
    """
    from app.data_persistence import export_all_data
    import json
    
    data = export_all_data(db)
    json_str = json.dumps(data, indent=2, default=str)
    
    return Response(
        content=json_str,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=subtrack_backup_{date.today().isoformat()}.json"}
    )


@router.get("/export/data-string")
async def export_data_string(db: Session = Depends(get_db)):
    """
    Export all data as a base64-encoded string.
    
    IMPORTANT: Copy this string and set it as the SUBTRACK_DATA environment 
    variable in your Railway/Render/Heroku deployment settings.
    
    This ensures your data persists across deployments!
    """
    from app.data_persistence import get_data_as_base64, export_all_data
    
    data = export_all_data(db)
    base64_string = get_data_as_base64(db)
    
    return {
        "message": "Copy the 'data_string' value below and set it as SUBTRACK_DATA environment variable in your deployment platform",
        "instructions": [
            "1. Copy the entire 'data_string' value (without quotes)",
            "2. Go to your Railway/Render/Heroku dashboard",
            "3. Find Environment Variables settings",
            "4. Create or update SUBTRACK_DATA with this value",
            "5. Redeploy your application",
            "Your data will now persist across deployments!"
        ],
        "data_stats": {
            "categories": len(data.get("categories", [])),
            "groups": len(data.get("groups", [])),
            "customers": len(data.get("customers", [])),
            "subscriptions": len(data.get("subscriptions", [])),
            "exported_at": data.get("exported_at")
        },
        "data_string": base64_string
    }


@router.post("/import/data-string")
async def import_data_string(payload: dict, db: Session = Depends(get_db)):
    """
    Import data from a base64-encoded string.
    
    Send a POST request with: {"data_string": "your_base64_string_here"}
    
    WARNING: This will add data to your database. Existing records with 
    the same IDs will NOT be overwritten.
    """
    import base64
    import json
    from app.data_persistence import import_data_to_db
    
    data_string = payload.get("data_string")
    if not data_string:
        return {"error": "Missing 'data_string' in request body"}
    
    try:
        json_str = base64.b64decode(data_string).decode('utf-8')
        data = json.loads(json_str)
        
        success = import_data_to_db(db, data)
        
        if success:
            return {
                "message": "Data imported successfully",
                "imported": {
                    "categories": len(data.get("categories", [])),
                    "groups": len(data.get("groups", [])),
                    "customers": len(data.get("customers", [])),
                    "subscriptions": len(data.get("subscriptions", []))
                }
            }
        else:
            return {"error": "Import failed - check server logs for details"}
    except Exception as e:
        return {"error": f"Import failed: {str(e)}"}


@router.post("/import/data")
async def import_data_json(payload: dict, db: Session = Depends(get_db)):
    """
    Import data from a JSON object (from file upload).
    
    Send a POST request with the JSON data directly.
    Expected format: {"categories": [...], "customers": [...], "subscriptions": [...], ...}
    
    WARNING: This will add data to your database. Existing records with 
    the same IDs will NOT be overwritten.
    """
    from app.data_persistence import import_data_to_db
    
    # Check if payload has the expected structure
    if not any(key in payload for key in ["categories", "customers", "subscriptions", "groups"]):
        return {"error": "Invalid data format. Expected JSON with categories, customers, subscriptions, or groups."}
    
    try:
        success = import_data_to_db(db, payload)
        
        if success:
            return {
                "message": "Data imported successfully",
                "imported": {
                    "categories": len(payload.get("categories", [])),
                    "groups": len(payload.get("groups", [])),
                    "customers": len(payload.get("customers", [])),
                    "subscriptions": len(payload.get("subscriptions", []))
                }
            }
        else:
            return {"error": "Import failed - check server logs for details"}
    except Exception as e:
        return {"error": f"Import failed: {str(e)}"}
